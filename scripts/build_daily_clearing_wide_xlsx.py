#!/usr/bin/env python3
"""
日清算结果查询电厂侧 — 长表转宽表 + 充放电时段着色（xlsx）

数据源
------
`source_data/日清算结果查询电厂侧.csv`：两实体长表
  - 独立储能充电：按小时 01:00..24:00
  - 一期：按 15 分钟 00:15..24:00

宽表规则（与既有 `日清算结果查询电厂侧_副本.csv` 对齐）
----------------------------------------
- 以 15 分钟为索引（一期有数据的日期用一期列；仅有充电的日期仍输出 96 行，一期列留空）
- 独立储能充电列：按「报表时刻」= ceil(时刻分钟/60) 对应到 01:00..24:00 的充电小时行合并
- 列名：`独立储能充电_*` / `一期_*`

着色规则
--------
- 读取策略结果 CSV（默认 jan25 不跨日回测），按 `charge_start`~`charge_end`、`discharge_start`~`discharge_end`
  对「查询日期」匹配行的整行加浅绿（充电）或浅红（放电）；同日先充后放，窗口不重叠。

用法
----
    python scripts/build_daily_clearing_wide_xlsx.py
    python scripts/build_daily_clearing_wide_xlsx.py --strategy path/to/strategy_result_nodaycross.csv
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
SRC_CSV = ROOT / "source_data" / "日清算结果查询电厂侧.csv"
OUT_CSV = ROOT / "source_data" / "日清算结果查询电厂侧_副本.csv"
OUT_XLSX = ROOT / "source_data" / "日清算结果查询电厂侧_副本.xlsx"

CHG_ENTITY_SUB = "独立储能充电"
Q1_ENTITY_SUB = "#1期"
CHG_PREFIX = "充电_"
Q1_PREFIX = "一期_"

# 充电_实际用电量 > 10 → 红色；一期_计量电量 > 0 → 绿色
FILL_CHARGE    = "FFCDD2"   # 浅红：充电中（充电_实际用电量 > 10）
FILL_DISCHARGE = "C8E6C9"   # 浅绿：放电中（一期_计量电量 > 0）


def _slot_labels_96() -> list[str]:
    out: list[str] = []
    for h in range(24):
        out.append(f"{h:02d}:15")
        out.append(f"{h:02d}:30")
        out.append(f"{h:02d}:45")
        out.append("24:00" if h == 23 else f"{h + 1:02d}:00")
    return out


def _slot_to_report_time(slot: str) -> str:
    """15 分钟刻度 → 日清算充电行时刻标签（01:00..24:00）。"""
    parts = slot.split(":")
    h, m = int(parts[0]), int(parts[1])
    mins = h * 60 + m
    if mins <= 0:
        raise ValueError(slot)
    k = math.ceil(mins / 60)
    if k > 24:
        k = 24
    return "24:00" if k == 24 else f"{k:02d}:00"


def _slot_calendar_hour(slot: str) -> int:
    """用于与策略 charge_start..charge_end（整点小时）对齐。"""
    h, m = map(int, slot.split(":"))
    if h == 24:
        return 24
    return h


def _load_strategy(path: Path | None) -> pd.DataFrame | None:
    if path is None or not path.is_file():
        return None
    st = pd.read_csv(path)
    st["date"] = pd.to_datetime(st["date"]).dt.date
    return st


def _row_tag(st: pd.DataFrame | None, day, slot: str) -> str:
    if st is None:
        return ""
    row = st.loc[st["date"] == day]
    if row.empty:
        return ""
    r = row.iloc[0]
    h = _slot_calendar_hour(slot)
    cs, ce = int(r["charge_start"]), int(r["charge_end"])
    ds, de = int(r["discharge_start"]), int(r["discharge_end"])
    if cs <= h <= ce:
        return "充电"
    if ds <= h <= de:
        return "放电"
    return ""


def build_wide(raw: pd.DataFrame) -> pd.DataFrame:
    """
    与旧副本完全对齐的宽表构建规则：
    - 15列固定输出：与 日清算结果查询电厂侧_副本.csv 列名一致
    - 充电列：仅在每小时第一个 15min 槽（xx:15）填入数据，其余三槽留空
    - 一期列：每个 15min 槽独立填入
    - 实体编码列单独保留（不合并到其他列名）
    """
    chg = raw[raw["实体名称"].str.contains(CHG_ENTITY_SUB, na=False)].copy()
    q1  = raw[raw["实体名称"].str.contains(Q1_ENTITY_SUB,  na=False)].copy()

    # 充电选取的字段（去掉实体编码）
    CHG_FIELDS = ["实际用电量", "电能电价", "电能电费", "实时节点电价", "曲线合理度取均值"]
    # 一期选取的字段（去掉实体编码）
    Q1_FIELDS  = ["计量电量", "电能电价", "电能电费", "省内实时出清电力", "省内实时节点电价"]

    slots = _slot_labels_96()
    all_days = sorted(set(chg["查询日期"].unique()) | set(q1["查询日期"].unique()))

    # 预索引：每天的充电/一期数据字典
    def _chg_map(d):
        sub = chg[chg["查询日期"] == d]
        m = {}
        for _, rr in sub.iterrows():
            t = str(rr["时刻"])
            m[t] = {f: rr.get(f, pd.NA) for f in CHG_FIELDS}
        return m

    def _q1_map(d):
        sub = q1[q1["查询日期"] == d]
        m = {}
        for _, rr in sub.iterrows():
            t = str(rr["时刻"])
            m[t] = {f: rr.get(f, pd.NA) for f in Q1_FIELDS}
        return m

    # 判断某 15min 槽是否是该充电小时的"第一个槽"（xx:15）
    def _is_first_slot(slot: str) -> bool:
        return slot.endswith(":15") or slot == "24:00"

    rows = []
    for d in all_days:
        cm = _chg_map(d)
        qm = _q1_map(d)

        for slot in slots:
            rep = _slot_to_report_time(slot)
            rec = {"查询日期": d, "一期_时刻": slot, f"{CHG_PREFIX}报表时刻": rep}

            # 充电列：仅首槽填入，其余留空（与旧副本一致）
            if _is_first_slot(slot) and rep in cm:
                for f, v in cm[rep].items():
                    rec[f"{CHG_PREFIX}{f}"] = v
            else:
                for f in CHG_FIELDS:
                    rec[f"{CHG_PREFIX}{f}"] = pd.NA

            # 一期列：逐槽填入
            if slot in qm:
                for f, v in qm[slot].items():
                    rec[f"{Q1_PREFIX}{f}"] = v
            else:
                for f in Q1_FIELDS:
                    rec[f"{Q1_PREFIX}{f}"] = pd.NA

            rows.append(rec)

    wide = pd.DataFrame(rows)

    # 固定列顺序（去掉两列实体编码，充电前缀简化为"充电_"）
    col_order = [
        "查询日期",
        "一期_时刻",
        f"{CHG_PREFIX}报表时刻",
        f"{CHG_PREFIX}实际用电量",
        f"{CHG_PREFIX}电能电价",
        f"{CHG_PREFIX}电能电费",
        f"{CHG_PREFIX}实时节点电价",
        f"{CHG_PREFIX}曲线合理度取均值",
        f"{Q1_PREFIX}计量电量",
        f"{Q1_PREFIX}电能电价",
        f"{Q1_PREFIX}电能电费",
        f"{Q1_PREFIX}省内实时出清电力",
        f"{Q1_PREFIX}省内实时节点电价",
    ]
    wide = wide[[c for c in col_order if c in wide.columns]]
    return wide


def _apply_colors_xlsx(
    xlsx_path: Path,
    df: pd.DataFrame,
    strategy: pd.DataFrame | None = None,  # 不再使用，保留参数兼容
) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # 着色规则：直接读宽表数据列
    #   充电_实际用电量 > 10  → 红色（充电中）
    #   一期_计量电量   > 0   → 绿色（放电中）
    chg_col = df.get("充电_实际用电量", pd.Series(dtype=float))
    q1_col  = df.get("一期_计量电量",   pd.Series(dtype=float))

    chg_vals = pd.to_numeric(chg_col, errors="coerce").fillna(0)
    q1_vals  = pd.to_numeric(q1_col,  errors="coerce").fillna(0)

    tags = []
    for c, q in zip(chg_vals, q1_vals):
        if c > 10:
            tags.append("充电")
        elif q > 0:
            tags.append("放电")
        else:
            tags.append("")

    colored_rows = {i + 2: t for i, t in enumerate(tags) if t}
    if not colored_rows:
        print("  未找到需要着色的行（充电_实际用电量均≤10 且 一期_计量电量均≤0）")
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active
    fill_chg = PatternFill("solid", fgColor=FILL_CHARGE)      # 红：充电
    fill_dis = PatternFill("solid", fgColor=FILL_DISCHARGE)   # 绿：放电

    for ri, tag in colored_rows.items():
        fl = fill_chg if tag == "充电" else fill_dis
        for cell in ws[ri]:
            cell.fill = fl
    wb.save(xlsx_path)
    print(f"  着色行数: 充电红色={sum(1 for t in tags if t=='充电')}行"
          f"  放电绿色={sum(1 for t in tags if t=='放电')}行")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--src",
        type=Path,
        default=SRC_CSV,
        help="源 CSV（默认: 日清算结果查询电厂侧.csv）",
    )
    ap.add_argument(
        "--out-csv",
        type=Path,
        default=OUT_CSV,
        help="输出宽表 CSV 路径",
    )
    ap.add_argument(
        "--out-xlsx",
        type=Path,
        default=OUT_XLSX,
        help="输出宽表 XLSX 路径",
    )
    ap.add_argument(
        "--strategy",
        type=Path,
        default=ROOT / "output/experiments/v8.0-jan25-sudun500/strategy_result_nodaycross.csv",
        help="含 charge_start/charge_end/discharge_* 的策略 CSV；不存在则不着色",
    )
    ap.add_argument("--no-xlsx", action="store_true", help="只写 CSV 不写 xlsx")
    args = ap.parse_args()

    raw = pd.read_csv(args.src)
    wide = build_wide(raw)
    args.out_csv.parent.mkdir(parents=True, exist_ok=True)
    wide.to_csv(args.out_csv, index=False)

    strategy = _load_strategy(args.strategy)
    if strategy is not None:
        print(f"已加载策略着色: {args.strategy} ({len(strategy)} 天)")
    else:
        print("未找到策略文件，xlsx 不着色")

    if not args.no_xlsx:
        args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wide.to_excel(args.out_xlsx, index=False)
        _apply_colors_xlsx(args.out_xlsx, wide, strategy)

    print(f"宽表 CSV : {args.out_csv}  ({len(wide)} 行)")
    if not args.no_xlsx:
        print(f"宽表 XLSX: {args.out_xlsx}")


if __name__ == "__main__":
    main()
